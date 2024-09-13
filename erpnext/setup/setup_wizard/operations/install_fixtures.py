# Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# License: GNU General Public License v3. See license.txt


import json
import os
from pathlib import Path

import frappe
from frappe import _
from frappe.desk.doctype.global_search_settings.global_search_settings import (
	update_global_search_doctypes,
)
from frappe.desk.page.setup_wizard.setup_wizard import make_records
from frappe.utils import cstr, getdate

from erpnext.accounts.doctype.account.account import RootNotEditable
from erpnext.regional.address_template.setup import set_up_address_templates


def read_lines(filename: str) -> list[str]:
	"""Return a list of lines from a file in the data directory."""
	return (Path(__file__).parent.parent / "data" / filename).read_text().splitlines()


def install(country=None):
	records = [
		# ensure at least an empty Address Template exists for this Country
		{"doctype": "Address Template", "country": country},
		# item group
		{
			"doctype": "Item Group",
			"item_group_name": _("All Item Groups"),
			"is_group": 1,
			"parent_item_group": "",
		},
		{
			"doctype": "Item Group",
			"item_group_name": _("Products"),
			"is_group": 0,
			"parent_item_group": _("All Item Groups"),
			"show_in_website": 1,
		},
		# Stock Entry Type
		{"doctype": "Stock Entry Type", "name": "Material Issue", "purpose": "Material Issue"},
		{"doctype": "Stock Entry Type", "name": "Material Receipt", "purpose": "Material Receipt"},
		{
			"doctype": "Stock Entry Type",
			"name": "Material Transfer",
			"purpose": "Material Transfer",
		},
		{"doctype": "Stock Entry Type", "name": "Manufacture", "purpose": "Manufacture"},
		{"doctype": "Stock Entry Type", "name": "Repack", "purpose": "Repack"},
		{
			"doctype": "Stock Entry Type",
			"name": "Send to Subcontractor",
			"purpose": "Send to Subcontractor",
		},
		{
			"doctype": "Stock Entry Type",
			"name": "Material Transfer for Manufacture",
			"purpose": "Material Transfer for Manufacture",
		},
		{
			"doctype": "Stock Entry Type",
			"name": "Material Consumption for Manufacture",
			"purpose": "Material Consumption for Manufacture",
		},
		
		# Mode of Payment
		{
			"doctype": "Mode of Payment",
			"mode_of_payment": "Check" if country == "United States" else _("Cheque"),
			"type": "Bank",
		},
		{"doctype": "Mode of Payment", "mode_of_payment": _("Cash"), "type": "Cash"},
		{"doctype": "Mode of Payment", "mode_of_payment": _("Credit Card"), "type": "Bank"},
		{"doctype": "Mode of Payment", "mode_of_payment": _("Wire Transfer"), "type": "Bank"},
		{"doctype": "Mode of Payment", "mode_of_payment": _("Bank Draft"), "type": "Bank"},
		# Issue Priority
		{"doctype": "Issue Priority", "name": _("Low")},
		{"doctype": "Issue Priority", "name": _("Medium")},
		{"doctype": "Issue Priority", "name": _("High")},
		# Party Type
		{"doctype": "Party Type", "party_type": "Customer", "account_type": "Receivable"},
		{"doctype": "Party Type", "party_type": "Supplier", "account_type": "Payable"},
		{"doctype": "Party Type", "party_type": "Employee", "account_type": "Payable"},
		{"doctype": "Party Type", "party_type": "Shareholder", "account_type": "Payable"},
		# Opportunity Type
		{"doctype": "Opportunity Type", "name": _("Sales")},
		{"doctype": "Opportunity Type", "name": _("Support")},
		{"doctype": "Opportunity Type", "name": _("Maintenance")},
		# Project Type
		{"doctype": "Project Type", "project_type": "Internal"},
		{"doctype": "Project Type", "project_type": "External"},
		{"doctype": "Project Type", "project_type": "Other"},
		# Print Heading
		{"doctype": "Print Heading", "print_heading": _("Credit Note")},
		{"doctype": "Print Heading", "print_heading": _("Debit Note")},
		# Share Management
		{"doctype": "Share Type", "title": _("Equity")},
		{"doctype": "Share Type", "title": _("Preference")},
		# Market Segments
		{"doctype": "Market Segment", "market_segment": _("Lower Income")},
		{"doctype": "Market Segment", "market_segment": _("Middle Income")},
		{"doctype": "Market Segment", "market_segment": _("Upper Income")},
		# Warehouse Type
		{"doctype": "Warehouse Type", "name": "Transit"},
	]

	for doctype, title_field, filename in (
		# ("Designation", "designation_name", "designation.txt"),
		("Sales Stage", "stage_name", "sales_stage.txt"),
		("Industry Type", "industry", "industry_type.txt"),
		# ("Lead Source", "source_name", "lead_source.txt"),
		("Sales Partner Type", "sales_partner_type", "sales_partner_type.txt"),
	):
		records += [{"doctype": doctype, title_field: title} for title in read_lines(filename)]

	base_path = frappe.get_app_path("erpnext", "stock", "doctype")
	response = frappe.read_file(os.path.join(base_path, "delivery_trip/dispatch_notification_template.html"))

	records += [
		{
			"doctype": "Email Template",
			"name": _("Dispatch Notification"),
			"response": response,
			"subject": _("Your order is out for delivery!"),
			"owner": frappe.session.user,
		}
	]

	# Records for the Supplier Scorecard
	from erpnext.buying.doctype.supplier_scorecard.supplier_scorecard import make_default_records

	make_default_records()
	make_records(records)
	set_up_address_templates(default_country=country)
	update_selling_defaults()
	update_buying_defaults()
	# add_uom_data()
	update_item_variant_settings()
	update_global_search_doctypes()


def update_selling_defaults():
	selling_settings = frappe.get_doc("Selling Settings")
	selling_settings.cust_master_name = "Customer Name"
	selling_settings.so_required = "No"
	selling_settings.dn_required = "No"
	selling_settings.allow_multiple_items = 1
	selling_settings.sales_update_frequency = "Each Transaction"
	selling_settings.save()


def update_buying_defaults():
	buying_settings = frappe.get_doc("Buying Settings")
	buying_settings.supp_master_name = "Supplier Name"
	buying_settings.po_required = "No"
	buying_settings.pr_required = "No"
	buying_settings.maintain_same_rate = 1
	buying_settings.allow_multiple_items = 1
	buying_settings.save()


def update_item_variant_settings():
	# set no copy fields of an item doctype to item variant settings
	doc = frappe.get_doc("Item Variant Settings")
	doc.set_default_fields()
	doc.save()


def add_uom_data():
	# add UOMs
	uoms = json.loads(
		open(frappe.get_app_path("erpnext", "setup", "setup_wizard", "data", "uom_data.json")).read()
	)
	for d in uoms:
		if not frappe.db.exists("UOM", _(d.get("uom_name"))):
			frappe.get_doc(
				{
					"doctype": "UOM",
					"uom_name": _(d.get("uom_name")),
					"name": _(d.get("uom_name")),
					"must_be_whole_number": d.get("must_be_whole_number"),
					"enabled": 1,
				}
			).db_insert()

	# bootstrap uom conversion factors
	uom_conversions = json.loads(
		open(
			frappe.get_app_path("erpnext", "setup", "setup_wizard", "data", "uom_conversion_data.json")
		).read()
	)
	for d in uom_conversions:
		if not frappe.db.exists("UOM Category", _(d.get("category"))):
			frappe.get_doc({"doctype": "UOM Category", "category_name": _(d.get("category"))}).db_insert()

		if not frappe.db.exists(
			"UOM Conversion Factor",
			{"from_uom": _(d.get("from_uom")), "to_uom": _(d.get("to_uom"))},
		):
			frappe.get_doc(
				{
					"doctype": "UOM Conversion Factor",
					"category": _(d.get("category")),
					"from_uom": _(d.get("from_uom")),
					"to_uom": _(d.get("to_uom")),
					"value": d.get("value"),
				}
			).db_insert()


def add_market_segments():
	records = [
		# Market Segments
		{"doctype": "Market Segment", "market_segment": _("Lower Income")},
		{"doctype": "Market Segment", "market_segment": _("Middle Income")},
		{"doctype": "Market Segment", "market_segment": _("Upper Income")},
	]

	make_records(records)


def add_sale_stages():
	# Sale Stages
	records = [
		{"doctype": "Sales Stage", "stage_name": _("Prospecting")},
		{"doctype": "Sales Stage", "stage_name": _("Qualification")},
		{"doctype": "Sales Stage", "stage_name": _("Needs Analysis")},
		{"doctype": "Sales Stage", "stage_name": _("Value Proposition")},
		{"doctype": "Sales Stage", "stage_name": _("Identifying Decision Makers")},
		{"doctype": "Sales Stage", "stage_name": _("Perception Analysis")},
		{"doctype": "Sales Stage", "stage_name": _("Proposal/Price Quote")},
		{"doctype": "Sales Stage", "stage_name": _("Negotiation/Review")},
	]
	for sales_stage in records:
		frappe.get_doc(sales_stage).db_insert()


def install_company(args):
	records = [
		# Fiscal Year
		{
			"doctype": "Fiscal Year",
			"year": get_fy_details(args.fy_start_date, args.fy_end_date),
			"year_start_date": args.fy_start_date,
			"year_end_date": args.fy_end_date,
		},
		# Company
		{
			"doctype": "Company",
			"company_name": args.company_name,
			"enable_perpetual_inventory": 1,
			"abbr": args.company_abbr,
			"default_currency": args.currency,
			"country": args.country,
			"create_chart_of_accounts_based_on": "Standard Template",
			"chart_of_accounts": args.chart_of_accounts,
			"domain": args.domain,
		},
	]

	make_records(records)


def install_defaults(args=None):  # nosemgrep
	# enable default currency
	frappe.db.set_value("Currency", args.get("currency"), "enabled", 1)
	frappe.db.set_single_value("Stock Settings", "email_footer_address", args.get("company_name"))

	set_global_defaults(args)
	# update_stock_settings()

	args.update({"set_default": 1})

def read_excel(file):
	setup_wizard_path = frappe.get_module_path("setup", 'setup_wizard',"operations")
	file_path = frappe.get_site_path(setup_wizard_path,file)
	from openpyxl import load_workbook
	wb = load_workbook(filename=file_path)
	ws = wb.active  # Chọn sheet đầu tiên (active sheet)

	# Bước 3: Duyệt qua từng hàng trong sheet
	data = []
	for row in ws.iter_rows(values_only=True):
		# Bước 4: Lưu trữ dữ liệu của từng hàng
		data.append(list(row))
	final_data = [] 
	del data[0]

	for row in data:
		if not row[1]:
			row[1] = row[0]
			row[3] = row[2]
		final_data.append(row)

	return final_data

def install_accounting(args=None):
	# xóa dữ liệu mặc định accouting
	frappe.db.delete("Account")
	#đọc excel
	accouting_type = args.accouting_type
	company =  args.company_name
	company = frappe.get_doc("Company", company)
	data = read_excel(f"{accouting_type}.xlsx")
	from erpnext.accounts.doctype.chart_of_accounts_importer.chart_of_accounts_importer import build_forest,create_charts
	#xử lý dữ liệu excel vào doctype accouting
	frappe.local.flags.ignore_root_company_validation = True
	forest = build_forest(data)
	create_charts(company, custom_chart=forest, from_coa_importer=True)

	# trigger on_update for company to reset default accounts
	from erpnext.setup.doctype.company.company import install_country_fixtures
	update_accounting =  {
		"default_bank_account" : frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 1121}
				),
		"default_cash_account": frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 1111}
				),
		"default_receivable_account": frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 131}
				),
		"default_payable_account": frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 331}
				),
		"default_expense_account": frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 632}
				),
		"default_income_account": frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 5111}
				),
		"round_off_account": "",
		"default_deferred_revenue_account": frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 3387}
				),
		"accumulated_depreciation_account": frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 2141}
				),
		"depreciation_expense_account": frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 6424}
				),
		"default_employee_advance_account":frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 141}
				),
		"stock_adjustment_account": frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 811}
				),
		"stock_received_but_not_billed": frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 151}
				),
		"default_inventory_account":frappe.db.get_value(
				"Account", {"company": company.name, "account_number": 152}
			),
		"write_off_account": None,
		"exchange_gain_loss_account": None,
		"expenses_included_in_asset_valuation": None,
		"capital_work_in_progress_account": None,
		"asset_received_but_not_billed": None,
		"expenses_included_in_valuation": None,
		"disposal_account": None,
		"expenses_included_in_asset_valuation":None
	}
	# xử lý cấu hình accouting company
	if accouting_type == "tt133":
		update_accounting.update({
			"default_discount_account":frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 5111}
				),
			"default_payroll_payable_account":frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 334}
				) ,
			
		})
		company.update(
			update_accounting
		)
	elif accouting_type == "tt200" : 
		update_accounting.update({
				"default_deferred_expense_account": frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 142}
				),
				"default_discount_account": frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 5211}
				),
				"default_payroll_payable_account":frappe.db.get_value(
					"Account", {"company": company.name, "account_number": 3341}
				)})
	print("update_accounting====================",update_accounting)
	company.update(update_accounting)
	company.save()
	install_country_fixtures(company.name, company.country)
	company.create_default_tax_template()



	# thiết lập “Default valuation method” trong stock setting
	inven_valua_method = args.inven_valua_method
	print("inven_valua_method",inven_valua_method)
	stock_setting = frappe.get_doc("Stock Settings")
	stock_setting.set("valuation_method",inven_valua_method)
	stock_setting.save()

	# xử lý xóa dữ liệu không cần thiết
	frappe.db.delete("Supplier Group")
	frappe.db.delete("Sales Person")
	frappe.db.delete("Lead Source")
	frappe.db.delete("Customer Group")
	frappe.db.delete("Territory")
	frappe.db.delete("Designation")
	frappe.db.delete("Activity Type")
	if frappe.db.table_exists("Expense Claim Type"):
		frappe.db.delete("Expense Claim Type")
	if frappe.db.table_exists("Vehicle Service Item"):
		frappe.db.delete("Vehicle Service Item")
	frappe.db.delete("UOM")
	frappe.db.delete("Item Attribute")
	frappe.db.delete("UOM Conversion Factor")
	if frappe.db.table_exists("Item Group"):
		frappe.db.delete("Item Group")
	frappe.db.delete("Purchase Taxes and Charges Template")
	frappe.db.delete("Supplier Scorecard Variable")
	frappe.db.delete("Supplier Scorecard Standing")
	frappe.db.delete("Sales Taxes and Charges Template")
	frappe.db.delete("Department")
	frappe.db.delete("Energy Point Rule")
	if frappe.db.table_exists("Leave Type"):
		frappe.db.delete("Leave Type")
	if frappe.db.table_exists("Role Profile"):
		frappe.db.delete("Role Profile")
	frappe.db.delete("Warehouse")
	frappe.db.delete("Email Account")
	frappe.db.commit()
	# raise ValueError("vấn đề lúc xóa")


def set_global_defaults(args):
	global_defaults = frappe.get_doc("Global Defaults", "Global Defaults")

	global_defaults.update(
		{
			"default_currency": args.get("currency"),
			"default_company": args.get("company_name"),
			"country": args.get("country"),
		}
	)

	global_defaults.save()


def update_stock_settings():
	stock_settings = frappe.get_doc("Stock Settings")
	stock_settings.item_naming_by = "Item Code"
	stock_settings.valuation_method = "FIFO"
	# stock_settings.default_warehouse = frappe.db.get_value("Warehouse", {"warehouse_name": _("Stores")})
	# stock_settings.stock_uom = _("Nos")
	stock_settings.auto_indent = 1
	stock_settings.auto_insert_price_list_rate_if_missing = 1
	stock_settings.set_qty_in_transactions_based_on_serial_no_input = 1
	stock_settings.save()


def create_bank_account(args):
	if not args.get("bank_account"):
		args["bank_account"] = _("Bank Account")

	company_name = args.get("company_name")
	bank_account_group = frappe.db.get_value(
		"Account",
		{"account_type": "Bank", "is_group": 1, "root_type": "Asset", "company": company_name},
	)
	if bank_account_group:
		bank_account = frappe.get_doc(
			{
				"doctype": "Account",
				"account_name": args.get("bank_account"),
				"parent_account": bank_account_group,
				"is_group": 0,
				"company": company_name,
				"account_type": "Bank",
			}
		)
		try:
			doc = bank_account.insert()

			if args.get("set_default"):
				frappe.db.set_value(
					"Company",
					args.get("company_name"),
					"default_bank_account",
					bank_account.name,
					update_modified=False,
				)

			return doc

		except RootNotEditable:
			frappe.throw(_("Bank account cannot be named as {0}").format(args.get("bank_account")))
		except frappe.DuplicateEntryError:
			# bank account same as a CoA entry
			pass


def get_fy_details(fy_start_date, fy_end_date):
	start_year = getdate(fy_start_date).year
	if start_year == getdate(fy_end_date).year:
		fy = cstr(start_year)
	else:
		fy = cstr(start_year) + "-" + cstr(start_year + 1)
	return fy
